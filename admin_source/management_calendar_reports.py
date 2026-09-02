from jalali_utils import convert_dates_in_text, format_jalali, now_jalali, install_openpyxl_jalali_patch, install_pptx_jalali_patch
install_openpyxl_jalali_patch()
install_pptx_jalali_patch()
# -*- coding: utf-8 -*-
"""خروجی‌های دوره‌ای پایش اجرایی نسخه ۶.۶."""

def export_management_brief_excel(db, date_from, date_to, path, zone_id=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    brief = db.get_management_period_brief(date_from, date_to, zone_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "خلاصه مدیریتی"
    ws.sheet_view.rightToLeft = True
    ws.append(["گزارش دوره‌ای مدیریت محله‌محور", ""])
    ws.append(["از تاریخ", brief["date_from"]])
    ws.append(["تا تاریخ", brief["date_to"]])
    metrics = [
        ("مسائل ثبت‌شده", brief["issues_created"]),
        ("اقدامات ایجادشده", brief["actions_created"]),
        ("اقدامات تکمیل‌شده", brief["actions_completed"]),
        ("جلسات برگزارشده", brief["meetings_held"]),
        ("بازدیدهای میدانی", brief["field_visits"]),
        ("درخواست‌های مردمی", brief["citizen_requests"]),
        ("نامه‌های ثبت‌شده", brief["letters_registered"]),
        ("تأییدهای خاتمه‌یافته", brief["approvals_completed"]),
        ("رویدادهای تقویم", brief["calendar_events"]),
        ("مبلغ هزینه‌شده", brief["spent_amount"]),
        ("کل سررسیدها", brief["deadlines_total"]),
        ("سررسیدهای معوق", brief["overdue_deadlines"]),
    ]
    ws.append([])
    ws.append(["شاخص", "مقدار"])
    for row in metrics: ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="13294B")
    for cell in ws[5]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="13294B")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24

    detail = wb.create_sheet("سررسیدها")
    detail.sheet_view.rightToLeft = True
    headers = ["تاریخ", "دسته", "عنوان", "بلوک", "مسئول", "اولویت", "وضعیت", "روز باقیمانده"]
    detail.append(headers)
    for item in brief["deadlines"]:
        detail.append([item["date"], item["category"], item["title"], item["zone_name"],
                       item["responsible"], item["priority"], item["status"], item["days_remaining"]])
    for cell in detail[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="13294B")
    for row in detail.iter_rows():
        for cell in row: cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    widths = [14, 18, 44, 22, 22, 12, 18, 16]
    for idx, width in enumerate(widths, 1): detail.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)
    return path


def export_management_brief_pdf(db, date_from, date_to, path, zone_id=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from report_generator import _pdf_header_flowables, _make_pdf_table, _pdf_styles
    from pdf_text_utils import shape_fa

    brief = db.get_management_period_brief(date_from, date_to, zone_id)
    doc = SimpleDocTemplate(path, pagesize=A4, rightMargin=12*mm, leftMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    story = _pdf_header_flowables("گزارش دوره‌ای پایش اجرایی")
    _, _, section_style, normal_style = _pdf_styles()
    story.append(Paragraph(shape_fa(f"دوره گزارش: {brief['date_from']} تا {brief['date_to']}"), normal_style))
    story.append(Spacer(1, 4*mm))
    metric_rows = [
        ["مسائل ثبت‌شده", brief["issues_created"]], ["اقدامات ایجادشده", brief["actions_created"]],
        ["اقدامات تکمیل‌شده", brief["actions_completed"]], ["جلسات برگزارشده", brief["meetings_held"]],
        ["بازدیدهای میدانی", brief["field_visits"]], ["درخواست‌های مردمی", brief["citizen_requests"]],
        ["نامه‌های ثبت‌شده", brief["letters_registered"]], ["تأییدهای خاتمه‌یافته", brief["approvals_completed"]],
        ["رویدادهای تقویم", brief["calendar_events"]], ["سررسیدهای معوق", brief["overdue_deadlines"]],
    ]
    story.append(_make_pdf_table(["شاخص", "مقدار"], metric_rows, [115*mm, 45*mm]))
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(shape_fa("سررسیدها و برنامه‌های دوره"), section_style))
    deadline_rows = [[x["date"], x["category"], x["title"], x["zone_name"], x["responsible"], x["status"]]
                     for x in brief["deadlines"]]
    if deadline_rows:
        story.append(_make_pdf_table(["تاریخ", "نوع", "عنوان", "بلوک", "مسئول", "وضعیت"], deadline_rows,
                                     [23*mm, 28*mm, 52*mm, 28*mm, 28*mm, 24*mm]))
    else:
        story.append(Paragraph(shape_fa("در این دوره سررسیدی ثبت نشده است."), normal_style))
    doc.build(story)
    return path


