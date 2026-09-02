# -*- coding: utf-8 -*-
"""گردش تأیید چندمرحله‌ای، داشبورد تصمیم‌گیری و قالب‌های اداری نسخه ۶.۵."""

import os
from datetime import datetime

from PyQt5.QtCore import Qt, pyqtSignal, QUrl
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTabWidget, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView, QDialog, QDialogButtonBox,
    QFormLayout, QComboBox, QLineEdit, QTextEdit, QCheckBox, QMessageBox, QFileDialog,
    QGroupBox, QSplitter
)
from PyQt5.QtGui import QDesktopServices

from access_control import available_roles, role_title
from document_service import generate_document_from_template, template_fields
from header_widget import build_official_header
from jalali_widgets import JalaliDateEdit
from jalali_utils import convert_dates_in_text, iso_to_jalali, jalali_to_iso
from icon_manager import get_icon, set_button_style


ENTITY_LABELS = {
    "action": "اقدام اجرایی",
    "resolution": "مصوبه",
    "budget": "ردیف بودجه",
    "letter": "نامه اداری",
}
SOURCE_LABELS = {
    "none": "بدون رکورد مرتبط",
    "letter": "نامه اداری",
    "meeting": "جلسه شورای محله",
    "action": "اقدام اجرایی",
    "citizen_request": "درخواست مردمی",
}


def _table(headers, stretch=()):
    table = QTableWidget(0, len(headers))
    table.setHorizontalHeaderLabels(headers)
    table.setSelectionBehavior(QAbstractItemView.SelectRows)
    table.setSelectionMode(QAbstractItemView.SingleSelection)
    table.setEditTriggers(QAbstractItemView.NoEditTriggers)
    table.setAlternatingRowColors(True)
    table.verticalHeader().setVisible(False)
    for index in range(len(headers)):
        mode = QHeaderView.Stretch if index in stretch else QHeaderView.ResizeToContents
        table.horizontalHeader().setSectionResizeMode(index, mode)
    return table


class ApprovalRequestDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.entities = []
        self.setWindowTitle("ایجاد گردش تأیید")
        self.resize(650, 590)
        layout = QVBoxLayout(self)
        info = QLabel("رکورد موردنظر و مراحل تأیید را تعیین کنید. فقط یک گردش فعال برای هر رکورد مجاز است.")
        info.setWordWrap(True)
        layout.addWidget(info)
        form = QFormLayout()
        self.zone = QComboBox(); self.zone.addItem("همه بلوک‌ها", None)
        for item in db.get_zones(): self.zone.addItem(item["name"], item["id"])
        self.entity_type = QComboBox()
        for key, label in ENTITY_LABELS.items(): self.entity_type.addItem(label, key)
        self.entity = QComboBox()
        self.title = QLineEdit()
        self.title.setPlaceholderText("در صورت خالی بودن، عنوان رکورد استفاده می‌شود")
        self.due_date = JalaliDateEdit(); self.due_date.clear()
        self.notes = QTextEdit(); self.notes.setMaximumHeight(85)
        form.addRow("بلوک:", self.zone)
        form.addRow("نوع رکورد:", self.entity_type)
        form.addRow("رکورد:", self.entity)
        form.addRow("عنوان گردش:", self.title)
        form.addRow("مهلت تأیید:", self.due_date)
        form.addRow("توضیحات:", self.notes)
        layout.addLayout(form)

        steps_group = QGroupBox("مراحل تأیید")
        steps_form = QFormLayout(steps_group)
        self.level_count = QComboBox()
        self.level_count.addItem("یک مرحله", 1); self.level_count.addItem("دو مرحله", 2); self.level_count.addItem("سه مرحله", 3)
        self.level_count.setCurrentIndex(1)
        steps_form.addRow("تعداد مراحل:", self.level_count)
        role_items = [(key, title) for key, title in available_roles() if key in {"admin", "manager", "reporter", "gis", "field"}]
        self.step_roles = []
        defaults = ["manager", "admin", "admin"]
        for index in range(3):
            combo = QComboBox()
            for key, title in role_items: combo.addItem(title, key)
            pos = combo.findData(defaults[index])
            if pos >= 0: combo.setCurrentIndex(pos)
            self.step_roles.append(combo)
            steps_form.addRow(f"تأییدکننده مرحله {index + 1}:", combo)
        layout.addWidget(steps_group)

        self.zone.currentIndexChanged.connect(self.refresh_entities)
        self.entity_type.currentIndexChanged.connect(self.refresh_entities)
        self.entity.currentIndexChanged.connect(self._suggest_title)
        self.level_count.currentIndexChanged.connect(self._update_step_visibility)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Save).setText("ایجاد گردش تأیید")
        buttons.button(QDialogButtonBox.Cancel).setText("انصراف")
        buttons.accepted.connect(self._validate)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        self.refresh_entities()
        self._update_step_visibility()

    def _entity_rows(self):
        zone_id = self.zone.currentData()
        kind = self.entity_type.currentData()
        if kind == "action":
            return self.db.get_neighborhood_actions(zone_id)
        if kind == "resolution":
            return self.db.get_neighborhood_resolutions(zone_id=zone_id)
        if kind == "budget":
            return self.db.get_neighborhood_budgets(zone_id)
        if kind == "letter":
            return self.db.get_correspondence_letters(zone_id=zone_id, limit=5000)
        return []

    def refresh_entities(self):
        self.entities = self._entity_rows()
        self.entity.clear()
        for item in self.entities:
            title = item.get("title") or item.get("subject") or item.get("letter_number") or str(item.get("id"))
            approval = item.get("approval_status") or "نیاز ندارد"
            self.entity.addItem(f"{title} — {approval}", item.get("id"))
        self._suggest_title()

    def _suggest_title(self):
        row = self.entity.currentIndex()
        if row < 0 or row >= len(self.entities):
            return
        item = self.entities[row]
        title = item.get("title") or item.get("subject") or ""
        if not self.title.text().strip():
            self.title.setText(f"تأیید {ENTITY_LABELS.get(self.entity_type.currentData(), '')}: {title}")

    def _update_step_visibility(self):
        count = int(self.level_count.currentData() or 1)
        for index, combo in enumerate(self.step_roles):
            combo.setEnabled(index < count)

    def _validate(self):
        if self.entity.currentData() is None:
            QMessageBox.warning(self, "اطلاعات ناقص", "رکوردی برای تأیید انتخاب نشده است.")
            return
        self.accept()

    def values(self):
        count = int(self.level_count.currentData() or 1)
        steps = []
        for index in range(count):
            role = self.step_roles[index].currentData()
            steps.append({"approver_role": role, "approver_name": role_title(role)})
        row = self.entity.currentIndex()
        item = self.entities[row] if 0 <= row < len(self.entities) else {}
        return {
            "entity_type": self.entity_type.currentData(),
            "entity_id": self.entity.currentData(),
            "zone_id": item.get("zone_id") or self.zone.currentData(),
            "title": self.title.text().strip(),
            "due_date": self.due_date.isoDate() if self.due_date.text().strip() else None,
            "notes": self.notes.toPlainText().strip(),
            "steps": steps,
        }


class ApprovalDetailDialog(QDialog):
    def __init__(self, db, approval, parent=None):
        super().__init__(parent)
        self.setWindowTitle("جزئیات گردش تأیید")
        self.resize(760, 480)
        layout = QVBoxLayout(self)
        info = QLabel(
            f"<b>{approval.get('title')}</b><br>وضعیت: {approval.get('status')} — "
            f"مرحله {approval.get('current_step')} از {approval.get('total_steps')} — "
            f"بلوک: {approval.get('zone_name') or '—'}"
        )
        info.setWordWrap(True); layout.addWidget(info)
        table = _table(["مرحله", "تأییدکننده", "وضعیت", "تصمیم‌گیرنده", "زمان", "توضیح"], stretch=(1,5))
        steps = approval.get("steps") or db.get_approval_steps(approval["id"])
        table.setRowCount(len(steps))
        for row, step in enumerate(steps):
            approver = step.get("approver_user_name") or step.get("approver_name") or role_title(step.get("approver_role"))
            values = [step.get("step_order"), approver, step.get("status"), step.get("decided_by_name") or "—",
                      step.get("decided_at") or "—", step.get("decision_comment") or "—"]
            for col, value in enumerate(values): table.setItem(row, col, QTableWidgetItem(convert_dates_in_text(str(value))))
        layout.addWidget(table)
        notes = QLabel(f"توضیحات: {approval.get('notes') or '—'}")
        notes.setWordWrap(True); layout.addWidget(notes)
        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.button(QDialogButtonBox.Close).setText("بستن")
        buttons.rejected.connect(self.reject); buttons.accepted.connect(self.accept)
        layout.addWidget(buttons)


class DecisionDialog(QDialog):
    def __init__(self, approved=True, parent=None):
        super().__init__(parent)
        self.approved = approved
        self.setWindowTitle("تأیید مرحله" if approved else "رد مرحله")
        self.resize(500, 280)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("توضیح تصمیم را وارد کنید:"))
        self.comment = QTextEdit(); layout.addWidget(self.comment)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("ثبت تصمیم")
        buttons.button(QDialogButtonBox.Cancel).setText("انصراف")
        buttons.accepted.connect(self.accept); buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)


class TemplateDialog(QDialog):
    def __init__(self, db, template=None, parent=None):
        super().__init__(parent)
        self.db = db; self.template = template or {}
        self.setWindowTitle("قالب سند اداری")
        self.resize(760, 620)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.name = QLineEdit(self.template.get("name") or "")
        self.kind = QComboBox(); self.kind.addItems(db.DOCUMENT_TEMPLATE_TYPES)
        self.kind.setCurrentText(self.template.get("template_type") or db.DOCUMENT_TEMPLATE_TYPES[0])
        self.subject = QLineEdit(self.template.get("subject_template") or "")
        self.body = QTextEdit(self.template.get("body_template") or ""); self.body.setMinimumHeight(280)
        self.active = QCheckBox("قالب فعال باشد"); self.active.setChecked(bool(self.template.get("is_active", True)))
        form.addRow("نام قالب:", self.name); form.addRow("نوع:", self.kind); form.addRow("موضوع قالب:", self.subject)
        form.addRow("متن قالب:", self.body); form.addRow("", self.active)
        layout.addLayout(form)
        hint = QLabel("متغیرهای قابل استفاده: {zone_name}، {date}، {subject}، {due_date}، {responsible_office}، "
                      "{responsible_person}، {meeting_title}، {action_title}، {tracking_code} و سایر متغیرهای فرم تولید سند.")
        hint.setWordWrap(True); hint.setStyleSheet("color:#647184;"); layout.addWidget(hint)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Save).setText("ذخیره قالب")
        buttons.accepted.connect(self._validate); buttons.rejected.connect(self.reject); layout.addWidget(buttons)

    def _validate(self):
        if not self.name.text().strip() or not self.body.toPlainText().strip():
            QMessageBox.warning(self, "اطلاعات ناقص", "نام و متن قالب الزامی است.")
            return
        self.accept()

    def values(self):
        return {"name": self.name.text().strip(), "template_type": self.kind.currentText(),
                "subject_template": self.subject.text(), "body_template": self.body.toPlainText(),
                "is_active": self.active.isChecked()}


class GenerateDocumentDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db; self.source_rows = []
        self.setWindowTitle("تولید سند از قالب")
        self.resize(720, 690)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        self.template = QComboBox()
        for item in db.get_document_templates(active_only=True):
            self.template.addItem(f"{item['template_type']} — {item['name']}", item["id"])
        self.zone = QComboBox(); self.zone.addItem("بدون بلوک", None)
        for zone in db.get_zones(): self.zone.addItem(zone["name"], zone["id"])
        self.source_type = QComboBox()
        for key, label in SOURCE_LABELS.items(): self.source_type.addItem(label, key)
        self.source = QComboBox()
        self.output_format = QComboBox()
        self.output_format.addItem("Word", "docx")
        self.output_format.addItem("PDF", "pdf")
        self.output_format.addItem("Word و PDF", "both")
        self.doc_number = QLineEdit(); self.doc_number.setPlaceholderText("از نامه انتخاب‌شده خودکار پر می‌شود")
        self.doc_date = JalaliDateEdit(); self.doc_date.clear()
        self.recipient = QLineEdit(); self.recipient.setPlaceholderText("گیرنده نامه")
        self.sender = QLineEdit(); self.sender.setPlaceholderText("فرستنده نامه")
        self.subject = QLineEdit(); self.subject.setPlaceholderText("موضوع سند؛ از رکورد مرتبط خودکار پر می‌شود")
        self.due_date = JalaliDateEdit(); self.due_date.clear()
        self.responsible_person = QLineEdit(); self.responsible_office = QLineEdit()
        self.result = QTextEdit(); self.result.setMaximumHeight(100)
        form.addRow("قالب:", self.template); form.addRow("نوع خروجی:", self.output_format)
        form.addRow("بلوک:", self.zone)
        form.addRow("نوع رکورد مرتبط:", self.source_type); form.addRow("رکورد مرتبط:", self.source)
        form.addRow("شماره سند:", self.doc_number); form.addRow("تاریخ سند:", self.doc_date)
        form.addRow("گیرنده:", self.recipient); form.addRow("فرستنده:", self.sender)
        form.addRow("موضوع:", self.subject); form.addRow("مهلت:", self.due_date)
        form.addRow("مسئول پیگیری:", self.responsible_person); form.addRow("دستگاه مسئول:", self.responsible_office)
        form.addRow("نتیجه / توضیح تکمیلی:", self.result)
        layout.addLayout(form)
        self.fields_label = QLabel("—"); self.fields_label.setWordWrap(True); self.fields_label.setStyleSheet("color:#647184;")
        layout.addWidget(self.fields_label)
        self.zone.currentIndexChanged.connect(self.refresh_sources)
        self.source_type.currentIndexChanged.connect(self.refresh_sources)
        self.source.currentIndexChanged.connect(self.populate_from_source)
        self.template.currentIndexChanged.connect(self.refresh_fields)
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Save).setText("انتخاب مسیر و تولید سند")
        buttons.accepted.connect(self._validate); buttons.rejected.connect(self.reject); layout.addWidget(buttons)
        self.refresh_sources(); self.refresh_fields()

    def _rows(self):
        zone_id = self.zone.currentData(); kind = self.source_type.currentData()
        if kind == "letter": return self.db.get_correspondence_letters(zone_id=zone_id, limit=5000)
        if kind == "meeting": return self.db.get_neighborhood_meetings(zone_id)
        if kind == "action": return self.db.get_neighborhood_actions(zone_id)
        if kind == "citizen_request": return self.db.get_citizen_requests(zone_id)
        return []

    def refresh_sources(self):
        self.source_rows = self._rows(); self.source.clear()
        if self.source_type.currentData() == "none":
            self.source.addItem("بدون رکورد", None); return
        for item in self.source_rows:
            title = item.get("subject") or item.get("title") or item.get("tracking_code") or str(item.get("id"))
            self.source.addItem(title, item.get("id"))
        self.populate_from_source()

    def populate_from_source(self):
        kind = self.source_type.currentData()
        source_id = self.source.currentData()
        if kind == "none" or source_id is None:
            return
        item = None
        if kind == "letter":
            item = self.db.get_correspondence_letter(source_id)
        else:
            row = self.source.currentIndex()
            if 0 <= row < len(self.source_rows):
                item = self.source_rows[row]
        if not item:
            return
        self.subject.setText(str(item.get("subject") or item.get("title") or ""))
        self.doc_number.setText(str(item.get("letter_number") or item.get("tracking_code") or ""))
        self.doc_date.setText(iso_to_jalali(item.get("letter_date") or item.get("received_date") or item.get("meeting_date") or ""))
        self.recipient.setText(str(item.get("recipient") or item.get("responsible_office") or item.get("assigned_office") or ""))
        self.sender.setText(str(item.get("sender") or ""))
        self.due_date.setText(iso_to_jalali(item.get("due_date") or item.get("planned_end") or ""))
        if kind == "letter" and item.get("description"):
            self.result.setPlainText(str(item.get("description")))

    def refresh_fields(self):
        template = self.db.get_document_template(self.template.currentData()) if self.template.currentData() else None
        if not template: self.fields_label.setText("قالبی انتخاب نشده است."); return
        fields = template_fields(template.get("subject_template"), template.get("body_template"))
        self.fields_label.setText("متغیرهای این قالب: " + ("، ".join("{" + x + "}" for x in fields) if fields else "بدون متغیر"))

    def _validate(self):
        if self.template.currentData() is None:
            QMessageBox.warning(self, "اطلاعات ناقص", "قالبی برای تولید سند وجود ندارد."); return
        self.accept()

    def values(self):
        kind = self.source_type.currentData(); source_id = self.source.currentData()
        extra = {}
        field_values = {
            "subject": self.subject.text().strip(), "due_date": (jalali_to_iso(self.due_date.text().strip()) if self.due_date.text().strip() else ""),
            "responsible_person": self.responsible_person.text().strip(),
            "responsible_office": self.responsible_office.text().strip(),
            "result": self.result.toPlainText().strip(),
            "letter_number": self.doc_number.text().strip(), "letter_date": (jalali_to_iso(self.doc_date.text().strip()) if self.doc_date.text().strip() else ""),
            "recipient": self.recipient.text().strip(), "sender": self.sender.text().strip(),
        }
        extra.update({key: value for key, value in field_values.items() if value})
        kwargs = {"zone_id": self.zone.currentData(), "extra": extra}
        if kind == "letter": kwargs["letter_id"] = source_id
        elif kind == "meeting": kwargs["meeting_id"] = source_id
        elif kind == "action": kwargs["action_id"] = source_id
        elif kind == "citizen_request": kwargs["citizen_request_id"] = source_id
        metadata = {
            "number": self.doc_number.text().strip(), "date": self.doc_date.text().strip(),
            "recipient": self.recipient.text().strip(), "sender": self.sender.text().strip(),
        }
        return {"template_id": self.template.currentData(), "source_type": kind,
                "source_id": source_id, "output_format": self.output_format.currentData(),
                "metadata": {k: v for k, v in metadata.items() if v}, **kwargs}


class ApprovalTemplatesWindow(QWidget):
    back_requested = pyqtSignal()

    def __init__(self, db):
        super().__init__()
        self.db = db
        self.current_user = db.get_current_user() or {}
        self.can_manage = self.current_user.get("role") in {"admin", "manager"}
        self.approvals = []; self.templates = []; self.documents = []; self.kpi_rows = []
        self.setWindowTitle("گردش تأیید و اسناد اداری")
        self.resize(1480, 900)
        self._build_ui(); self.refresh()

    def _build_ui(self):
        root = QVBoxLayout(self); root.setContentsMargins(0,0,0,0); root.setSpacing(0)
        root.addWidget(build_official_header("گردش تأیید، اسناد استاندارد و داشبورد تصمیم‌گیری", self.db))
        bar_widget = QWidget(); bar = QHBoxLayout(bar_widget); bar.setContentsMargins(18,8,18,8)
        back = QPushButton("بازگشت به داشبورد"); set_button_style(back, "back", "ghost"); back.clicked.connect(self.back_requested.emit)
        refresh = QPushButton("بروزرسانی"); set_button_style(refresh, "refresh", "secondary"); refresh.clicked.connect(self.refresh)
        bar.addWidget(back); bar.addWidget(refresh); bar.addStretch()
        self.summary = QLabel("—"); self.summary.setStyleSheet("font-weight:800; color:#13294b;"); bar.addWidget(self.summary)
        root.addWidget(bar_widget)
        self.tabs = QTabWidget(); root.addWidget(self.tabs, 1)
        self._build_approval_tab(); self._build_kpi_tab(); self._build_templates_tab(); self._build_documents_tab()

    def _build_approval_tab(self):
        page = QWidget(); layout = QVBoxLayout(page); layout.setContentsMargins(14,14,14,14)
        filters = QHBoxLayout()
        self.approval_status = QComboBox(); self.approval_status.addItem("همه وضعیت‌ها", None)
        self.approval_status.addItems(self.db.APPROVAL_STATUSES); self.approval_status.currentIndexChanged.connect(self.refresh_approvals)
        self.approval_zone = QComboBox(); self.approval_zone.addItem("همه بلوک‌ها", None)
        for z in self.db.get_zones(): self.approval_zone.addItem(z["name"], z["id"])
        self.approval_zone.currentIndexChanged.connect(self.refresh_approvals)
        self.only_mine = QCheckBox("فقط کارتابل من"); self.only_mine.stateChanged.connect(self.refresh_approvals)
        filters.addWidget(self.approval_status); filters.addWidget(self.approval_zone); filters.addWidget(self.only_mine); filters.addStretch()
        new_btn = QPushButton("ایجاد گردش تأیید"); set_button_style(new_btn, "plus", "success"); new_btn.clicked.connect(self.new_approval)
        new_btn.setEnabled(self.can_manage); filters.addWidget(new_btn); layout.addLayout(filters)
        self.approval_table = _table(["شناسه", "عنوان", "نوع", "بلوک", "مرحله", "مهلت", "وضعیت", "درخواست‌کننده"], stretch=(1,3))
        self.approval_table.doubleClicked.connect(self.show_approval_detail); layout.addWidget(self.approval_table, 1)
        actions = QHBoxLayout()
        detail = QPushButton("جزئیات مراحل"); detail.clicked.connect(self.show_approval_detail); actions.addWidget(detail)
        approve = QPushButton("تأیید مرحله"); set_button_style(approve, "check", "success"); approve.clicked.connect(lambda: self.decide(True)); actions.addWidget(approve)
        reject = QPushButton("رد مرحله"); set_button_style(reject, "close", "danger"); reject.clicked.connect(lambda: self.decide(False)); actions.addWidget(reject)
        cancel = QPushButton("لغو گردش"); cancel.clicked.connect(self.cancel_approval); cancel.setEnabled(self.can_manage); actions.addWidget(cancel)
        actions.addStretch(); layout.addLayout(actions)
        self.tabs.addTab(page, get_icon("check", "navy"), "گردش تأیید")

    def _build_kpi_tab(self):
        page = QWidget(); layout = QVBoxLayout(page); layout.setContentsMargins(14,14,14,14)
        info = QLabel("بلوک‌ها بر اساس امتیاز عملکرد و ریسک‌های باز مرتب شده‌اند؛ ردیف‌های پایین‌تر نیازمند توجه مدیریتی بیشتری هستند.")
        info.setWordWrap(True); layout.addWidget(info)
        self.kpi_table = _table(["بلوک", "امتیاز", "سطح", "خانوار", "مسئله باز", "بحرانی", "اقدام فعال", "اقدام معوق", "مصوبه باز", "درخواست مردمی", "تأیید باز", "اضافه‌هزینه"], stretch=(0,))
        layout.addWidget(self.kpi_table, 1)
        row = QHBoxLayout(); export = QPushButton("خروجی Excel داشبورد"); set_button_style(export, "sheet", "primary"); export.clicked.connect(self.export_kpi_excel)
        row.addWidget(export); row.addStretch(); layout.addLayout(row)
        self.tabs.addTab(page, get_icon("report", "navy"), "داشبورد تصمیم‌گیری")

    def _build_templates_tab(self):
        page = QWidget(); layout = QVBoxLayout(page); layout.setContentsMargins(14,14,14,14)
        self.template_table = _table(["شناسه", "نام قالب", "نوع", "وضعیت", "موضوع", "آخرین تغییر"], stretch=(1,4))
        self.template_table.doubleClicked.connect(self.edit_template); layout.addWidget(self.template_table, 1)
        row = QHBoxLayout()
        add = QPushButton("قالب جدید"); add.clicked.connect(self.add_template); add.setEnabled(self.can_manage); row.addWidget(add)
        edit = QPushButton("ویرایش قالب"); edit.clicked.connect(self.edit_template); edit.setEnabled(self.can_manage); row.addWidget(edit)
        delete = QPushButton("حذف قالب"); delete.clicked.connect(self.delete_template); delete.setEnabled(self.can_manage); row.addWidget(delete)
        generate = QPushButton("تولید Word / PDF"); set_button_style(generate, "file", "primary"); generate.clicked.connect(self.generate_document); row.addWidget(generate)
        row.addStretch(); layout.addLayout(row)
        self.tabs.addTab(page, get_icon("file", "navy"), "قالب‌های اداری")

    def _build_documents_tab(self):
        page = QWidget(); layout = QVBoxLayout(page); layout.setContentsMargins(14,14,14,14)
        self.document_table = _table(["شناسه", "عنوان", "قالب", "بلوک", "رکورد مرتبط", "تولیدکننده", "تاریخ", "مسیر فایل"], stretch=(1,7))
        self.document_table.doubleClicked.connect(self.open_document); layout.addWidget(self.document_table, 1)
        row = QHBoxLayout(); open_btn = QPushButton("باز کردن سند"); open_btn.clicked.connect(self.open_document); row.addWidget(open_btn); row.addStretch(); layout.addLayout(row)
        self.tabs.addTab(page, get_icon("file", "navy"), "اسناد تولیدشده")

    def refresh(self):
        self.refresh_approvals(); self.refresh_kpis(); self.refresh_templates(); self.refresh_documents()
        stats = self.db.get_approval_stats()
        self.summary.setText(f"در انتظار تأیید: {stats['approvals_pending']} | کارتابل من: {stats['approvals_assigned_to_me']} | معوق: {stats['approvals_overdue']}")

    def refresh_approvals(self):
        if not hasattr(self, "approval_table"): return
        self.approvals = self.db.get_approval_requests(
            status=self.approval_status.currentData(), zone_id=self.approval_zone.currentData(),
            assigned_to_current=self.only_mine.isChecked(), limit=5000)
        self.approval_table.setRowCount(len(self.approvals))
        for row, item in enumerate(self.approvals):
            values = [item["id"], item["title"], ENTITY_LABELS.get(item["entity_type"], item["entity_type"]),
                      item.get("zone_name") or "—", f"{item['current_step']}/{item['total_steps']}",
                      item.get("due_date") or "—", item["status"], item.get("requested_by_name") or "system"]
            for col, value in enumerate(values): self.approval_table.setItem(row, col, QTableWidgetItem(convert_dates_in_text(str(value))))
            if item["status"] == "ردشده":
                for col in range(self.approval_table.columnCount()): self.approval_table.item(row, col).setBackground(QColor("#fde8e8"))
            elif item["status"] == "تأییدشده":
                for col in range(self.approval_table.columnCount()): self.approval_table.item(row, col).setBackground(QColor("#e8f5e9"))

    def _selected_approval(self):
        row = self.approval_table.currentRow()
        return self.approvals[row] if 0 <= row < len(self.approvals) else None

    def new_approval(self):
        dialog = ApprovalRequestDialog(self.db, self)
        if dialog.exec_() != QDialog.Accepted: return
        try:
            self.db.create_approval_request(**dialog.values()); self.refresh()
            QMessageBox.information(self, "ثبت شد", "گردش تأیید با موفقیت ایجاد شد.")
        except Exception as exc: QMessageBox.warning(self, "خطا", str(exc))

    def show_approval_detail(self):
        item = self._selected_approval()
        if not item: QMessageBox.information(self, "انتخاب", "ابتدا یک گردش را انتخاب کنید."); return
        ApprovalDetailDialog(self.db, self.db.get_approval_request(item["id"]), self).exec_()

    def decide(self, approved):
        item = self._selected_approval()
        if not item: QMessageBox.information(self, "انتخاب", "ابتدا یک گردش را انتخاب کنید."); return
        if not self.db.current_user_can_decide_approval(item["id"]):
            QMessageBox.warning(self, "عدم دسترسی", "این مرحله به کاربر یا نقش دیگری اختصاص دارد."); return
        dialog = DecisionDialog(approved, self)
        if dialog.exec_() != QDialog.Accepted: return
        try:
            self.db.decide_approval(item["id"], approved=approved, comment=dialog.comment.toPlainText().strip())
            self.refresh(); QMessageBox.information(self, "ثبت شد", "تصمیم این مرحله ثبت شد.")
        except Exception as exc: QMessageBox.warning(self, "خطا", str(exc))

    def cancel_approval(self):
        item = self._selected_approval()
        if not item: return
        if QMessageBox.question(self, "لغو گردش", "گردش تأیید انتخاب‌شده لغو شود؟", QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        try: self.db.cancel_approval(item["id"], "لغو از کارتابل"); self.refresh()
        except Exception as exc: QMessageBox.warning(self, "خطا", str(exc))

    def refresh_kpis(self):
        if not hasattr(self, "kpi_table"): return
        self.kpi_rows = self.db.get_zone_decision_rows(); self.kpi_table.setRowCount(len(self.kpi_rows))
        for row, item in enumerate(self.kpi_rows):
            values = [item["zone_name"], item["score"], item["level"], item["households"], item["open_issues"], item["critical_issues"],
                      item["active_actions"], item["overdue_actions"], item["pending_resolutions"], item["open_citizen_requests"],
                      item["pending_approvals"], item["budget_overruns"]]
            for col, value in enumerate(values): self.kpi_table.setItem(row, col, QTableWidgetItem(convert_dates_in_text(str(value))))
            color = "#fde8e8" if item["score"] < 50 else "#fff8db" if item["score"] < 70 else "#e8f5e9"
            for col in range(self.kpi_table.columnCount()): self.kpi_table.item(row, col).setBackground(QColor(color))

    def export_kpi_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "ذخیره داشبورد تصمیم‌گیری", "dashboard_decision.xlsx", "Excel (*.xlsx)")
        if not path: return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook(); ws = wb.active; ws.title = "داشبورد بلوک‌ها"; ws.sheet_view.rightToLeft = True
            headers = ["بلوک", "امتیاز", "سطح", "خانوار", "مسئله باز", "مسئله بحرانی", "اقدام فعال", "اقدام معوق", "مصوبه باز", "درخواست مردمی باز", "تأیید باز", "تأیید معوق", "بودجه تخصیصی", "بودجه مصرف‌شده", "اضافه‌هزینه"]
            ws.append(headers)
            for item in self.kpi_rows:
                ws.append([item["zone_name"], item["score"], item["level"], item["households"], item["open_issues"], item["critical_issues"],
                           item["active_actions"], item["overdue_actions"], item["pending_resolutions"], item["open_citizen_requests"],
                           item["pending_approvals"], item["overdue_approvals"], item["budget_allocated"], item["budget_spent"], item["budget_overruns"]])
            navy = PatternFill("solid", fgColor="13294B"); white = Font(color="FFFFFF", bold=True)
            thin = Side(style="thin", color="D7DBE3")
            for cell in ws[1]: cell.fill = navy; cell.font = white; cell.alignment = Alignment(horizontal="center")
            for row in ws.iter_rows(min_row=2):
                score = float(row[1].value or 0); fill = "FDE8E8" if score < 50 else "FFF8DB" if score < 70 else "E8F5E9"
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill); cell.border = Border(left=thin,right=thin,top=thin,bottom=thin)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            widths = [24,10,16,12,12,14,12,12,12,18,12,12,18,18,14]
            for idx, width in enumerate(widths, start=1): ws.column_dimensions[chr(64+idx)].width = width
            ws.freeze_panes = "A2"; wb.save(path)
            QMessageBox.information(self, "ذخیره شد", "داشبورد تصمیم‌گیری با موفقیت ذخیره شد.")
        except Exception as exc: QMessageBox.warning(self, "خطا", str(exc))

    def refresh_templates(self):
        if not hasattr(self, "template_table"): return
        self.templates = self.db.get_document_templates(); self.template_table.setRowCount(len(self.templates))
        for row, item in enumerate(self.templates):
            values = [item["id"], item["name"], item["template_type"], "فعال" if item["is_active"] else "غیرفعال",
                      item.get("subject_template") or "—", item.get("updated_at") or item.get("created_at")]
            for col, value in enumerate(values): self.template_table.setItem(row, col, QTableWidgetItem(convert_dates_in_text(str(value))))

    def _selected_template(self):
        row = self.template_table.currentRow(); return self.templates[row] if 0 <= row < len(self.templates) else None

    def add_template(self):
        dialog = TemplateDialog(self.db, parent=self)
        if dialog.exec_() == QDialog.Accepted:
            try: self.db.add_document_template(**dialog.values()); self.refresh_templates()
            except Exception as exc: QMessageBox.warning(self, "خطا", str(exc))

    def edit_template(self):
        item = self._selected_template()
        if not item: QMessageBox.information(self, "انتخاب", "ابتدا یک قالب را انتخاب کنید."); return
        dialog = TemplateDialog(self.db, item, self)
        if dialog.exec_() == QDialog.Accepted:
            try: self.db.update_document_template(item["id"], **dialog.values()); self.refresh_templates()
            except Exception as exc: QMessageBox.warning(self, "خطا", str(exc))

    def delete_template(self):
        item = self._selected_template()
        if not item: return
        if QMessageBox.question(self, "حذف قالب", "قالب انتخاب‌شده حذف شود؟", QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes: return
        try: self.db.delete_document_template(item["id"]); self.refresh_templates()
        except Exception as exc: QMessageBox.warning(self, "خطا", str(exc))

    def generate_document(self):
        dialog = GenerateDocumentDialog(self.db, self)
        if dialog.exec_() != QDialog.Accepted: return
        values = dialog.values()
        output_format = values.pop("output_format")
        metadata = values.pop("metadata", {})
        default_name = "document"
        if output_format == "pdf":
            path, _ = QFileDialog.getSaveFileName(self, "ذخیره سند PDF", default_name + ".pdf", "PDF (*.pdf)")
            if not path: return
            paths = [path if path.lower().endswith(".pdf") else path + ".pdf"]
        elif output_format == "both":
            path, _ = QFileDialog.getSaveFileName(self, "انتخاب نام پایه سند", default_name + ".docx", "Word (*.docx)")
            if not path: return
            base = os.path.splitext(path)[0]
            paths = [base + ".docx", base + ".pdf"]
        else:
            path, _ = QFileDialog.getSaveFileName(self, "ذخیره سند Word", default_name + ".docx", "Word (*.docx)")
            if not path: return
            paths = [path if path.lower().endswith(".docx") else path + ".docx"]
        try:
            context = self.db.build_document_context(
                zone_id=values["zone_id"], letter_id=values.get("letter_id"), meeting_id=values.get("meeting_id"),
                action_id=values.get("action_id"), citizen_request_id=values.get("citizen_request_id"),
                extra=values.get("extra"))
            results = []
            for output_path in paths:
                result = generate_document_from_template(
                    self.db, values["template_id"], output_path, context=context, zone_id=values["zone_id"],
                    related_entity_type=values["source_type"] if values["source_type"] != "none" else None,
                    related_entity_id=values["source_id"],
                    metadata={"creator": context.get("user_full_name"), **metadata})
                results.append(result)
            self.refresh_documents()
            files = "\n".join(item["path"] for item in results)
            token = results[0].get("verification_token") if results else "—"
            QMessageBox.information(self, "سند تولید شد", f"فایل‌ها ذخیره شدند:\n{files}\n\nکد اعتبارسنجی: {token}")
        except Exception as exc: QMessageBox.warning(self, "خطا", str(exc))

    def refresh_documents(self):
        if not hasattr(self, "document_table"): return
        self.documents = self.db.get_generated_documents(limit=2000); self.document_table.setRowCount(len(self.documents))
        for row, item in enumerate(self.documents):
            related = f"{SOURCE_LABELS.get(item.get('related_entity_type'), item.get('related_entity_type') or '—')} {item.get('related_entity_id') or ''}".strip()
            values = [item["id"], item["title"], item.get("template_name") or "—", item.get("zone_name") or "—", related,
                      item.get("created_by_name") or "system", item.get("created_at") or "—", item.get("file_path") or "—"]
            for col, value in enumerate(values): self.document_table.setItem(row, col, QTableWidgetItem(convert_dates_in_text(str(value))))

    def _selected_document(self):
        row = self.document_table.currentRow(); return self.documents[row] if 0 <= row < len(self.documents) else None

    def open_document(self):
        item = self._selected_document()
        if not item: QMessageBox.information(self, "انتخاب", "ابتدا یک سند را انتخاب کنید."); return
        path = item.get("file_path")
        if not path or not os.path.exists(path): QMessageBox.warning(self, "فایل پیدا نشد", "فایل سند در مسیر ثبت‌شده وجود ندارد."); return
        QDesktopServices.openUrl(QUrl.fromLocalFile(path))
